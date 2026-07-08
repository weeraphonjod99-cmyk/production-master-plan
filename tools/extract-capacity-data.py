import json
from datetime import UTC, datetime
from pathlib import Path

import openpyxl


SOURCE_WORKBOOK = Path("Production Capacity.xlsx")
OUTPUT_FILE = Path("capacity-data.js")
SOURCE_URL = (
    "https://docs.google.com/spreadsheets/d/"
    "1eXby1xmCjhp_C8H_r7OC8JmnLu00WRYq/edit?gid=1003607569#gid=1003607569"
)


def compact(value):
    if value is None:
        return ""
    if isinstance(value, float) and value.is_integer():
        return str(int(value))
    return str(value).strip()


def number(value):
    if isinstance(value, (int, float)) and not isinstance(value, bool):
        return float(value)
    try:
        return float(str(value).replace(",", "").strip())
    except (TypeError, ValueError):
        return 0.0


def rounded(value):
    value = number(value)
    if not value:
        return 0
    return round(value, 3)


def find_header_row(worksheet):
    for row_index, row in enumerate(worksheet.iter_rows(values_only=True), start=1):
        if compact(row[0]).lower() == "item" and "part" in compact(row[1]).lower():
            return row_index
    return None


def extract_records():
    workbook = openpyxl.load_workbook(SOURCE_WORKBOOK, data_only=True)
    records = []

    for worksheet in workbook.worksheets:
        header_row = find_header_row(worksheet)
        if not header_row:
            continue

        for row_index, row in enumerate(
            worksheet.iter_rows(min_row=header_row + 1, values_only=True),
            start=header_row + 1,
        ):
            item = number(row[0] if len(row) > 0 else None)
            part_name = compact(row[1] if len(row) > 1 else None)
            part_no = compact(row[2] if len(row) > 2 else None)
            per_day_100 = rounded(row[15] if len(row) > 15 else None)
            per_day_85 = rounded(row[17] if len(row) > 17 else None)

            if item <= 0 or not (part_name or part_no) or not (per_day_100 or per_day_85):
                continue

            records.append(
                {
                    "machine": compact(worksheet.title),
                    "sourceRow": row_index,
                    "item": int(item) if item.is_integer() else item,
                    "partName": part_name,
                    "partNo": part_no,
                    "step": compact(row[3] if len(row) > 3 else None),
                    "speed10Min": rounded(row[4] if len(row) > 4 else None),
                    "piecesPerMinute": {
                        "100": rounded(row[6] if len(row) > 6 else None),
                        "90": rounded(row[8] if len(row) > 8 else None),
                        "85": rounded(row[10] if len(row) > 10 else None),
                    },
                    "perDay8Hours": {
                        "100": per_day_100,
                        "90": rounded(row[16] if len(row) > 16 else None),
                        "85": per_day_85,
                    },
                    "machineTon": compact(row[24] if len(row) > 24 else None),
                    "machineNo": compact(row[25] if len(row) > 25 else None),
                    "targetPerDay": rounded(row[27] if len(row) > 27 else None),
                }
            )

    return records


records = extract_records()
payload = {
    "sourceUrl": SOURCE_URL,
    "generatedAt": datetime.now(UTC).isoformat(timespec="seconds").replace("+00:00", "Z"),
    "records": records,
}

OUTPUT_FILE.write_text(
    "window.PRODUCTION_CAPACITY_DATA = "
    + json.dumps(payload, ensure_ascii=False, indent=2)
    + ";\n",
    encoding="utf-8",
)

print(
    json.dumps(
        {
            "records": len(records),
            "machines": len({record["machine"] for record in records}),
            "outputFile": str(OUTPUT_FILE),
        },
        ensure_ascii=False,
        indent=2,
    )
)
